from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path("reports/kona_pi_hat_generation_alignment")
SOURCE = BASE_DIR / "kona_pi_hat_generation_matrix_reconstructed.xlsx"
OUTPUT = BASE_DIR / "kona_pi_hat_pcsu_jonathan_handoff.xlsx"

SPECIAL_REVIEW_IDS = {
    "BI118": "Faux Ray is present in PI_HAT/staging but does not currently map to a usable Kona biopsied PCSU matrix record.",
    "BI_K511": "High PI_HAT with BI_K54; staging row has no catalog link.",
    "BI_K54": "High PI_HAT with unresolved BI_K511; mapped as Winona but should be checked as part of the duplicate/identity review.",
    "BI_K55": "High PI_HAT with BI_KM35 and BI_K27; mapped as Jana Ray but should be checked as part of the duplicate/identity review.",
    "BI_KM35": "High PI_HAT with BI_K55 and BI_K27; staging row has no catalog link.",
    "BI134": "High PI_HAT with unresolved BI_K52; mapped as Orion but should be checked as part of the paired identity review.",
    "BI_K52": "High PI_HAT with BI134; placeholder matrix row only, with no catalog/name/age.",
    "BI_K27": "High PI_HAT with BI_K55 and BI_KM35; mapped as Independence Ray but should be checked because of the BK6/BK7 crossover context.",
}


def main() -> None:
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    source = load_workbook(SOURCE, data_only=True, read_only=True)

    sequence_crosswalk = read_sheet(source, "Sequence crosswalk")
    all_pairs = read_sheet(source, "All PI_HAT pairs")
    strong_pairs = read_sheet(source, "Strong pairs")
    pcsu_wide = read_sheet(source, "PI_HAT PCSU matrix")
    pcsu_long = read_sheet(source, "PI_HAT PCSU long")
    proportions = read_sheet(source, "PCSU proportions")
    maturity_summary = read_sheet(source, "Maturity summary")

    sequence_by_id = {clean(row.get("piHatSampleId")): row for row in sequence_crosswalk}

    wb = Workbook()
    wb.remove(wb.active)

    add_readme(wb)
    write_sheet(wb, "Sequence crosswalk", build_sequence_rows(sequence_crosswalk))
    write_sheet(wb, "PI_HAT pair results", build_pair_rows(all_pairs))
    write_sheet(wb, "Strong PI_HAT pairs", build_pair_rows(strong_pairs))
    write_sheet(wb, "PCSU matrix wide", build_wide_rows(pcsu_wide))
    write_sheet(wb, "PCSU matrix long", build_long_rows(pcsu_long))
    write_sheet(wb, "Special review IDs", build_special_review_rows(strong_pairs, sequence_by_id))
    write_sheet(wb, "PCSU proportions", proportions)
    write_sheet(wb, "Maturity summary", maturity_summary)

    for ws in wb.worksheets:
        style_sheet(ws)
    highlight_special_review(wb["Sequence crosswalk"], "sequence_pk")
    highlight_special_review(wb["PI_HAT pair results"], "sequence_pk_a", "sequence_pk_b")
    highlight_special_review(wb["Strong PI_HAT pairs"], "sequence_pk_a", "sequence_pk_b")
    highlight_special_review(wb["PCSU matrix wide"], "row_sequence_pk")
    highlight_special_review(wb["PCSU matrix long"], "row_sequence_pk", "column_sequence_pk")

    wb.save(OUTPUT)
    print(OUTPUT)


def read_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def add_readme(wb: Workbook) -> None:
    rows = [
        {
            "item": "Purpose",
            "description": "Jonathan handoff workbook linking Kona PI_HAT genetic pair results to the Kona biopsied PCSU generation matrix.",
        },
        {
            "item": "Primary key",
            "description": "sequence_pk is the PI_HAT sample ID. It appears in the sequence crosswalk, PI_HAT pair sheets, and PCSU matrix sheets.",
        },
        {
            "item": "Pair key",
            "description": "pair_pk is a stable undirected key built from sorted sequence_pk values. directional_pair_pk preserves A-to-B order.",
        },
        {
            "item": "PCSU codes",
            "description": "P = row animal could be parent of column animal; C = row animal could be child of column animal; S = same-generation compatible; U = unknown/insufficient age evidence; dash = same animal; unmatched = no usable matrix mapping.",
        },
        {
            "item": "Maturity assumptions",
            "description": "The PCSU matrix in this workbook uses the midpoint Kona biopsied model: male maturity age 6.5 years, female maturity age 11.5 years. Low/high sensitivity outputs are retained in the analysis workbook.",
        },
        {
            "item": "Special review IDs",
            "description": "The Special review IDs sheet flags the eight sequence IDs from the strong unmatched/high-PI_HAT identity review cluster.",
        },
    ]
    write_sheet(wb, "README", rows)


def build_sequence_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        sequence_pk = clean(row.get("piHatSampleId"))
        output.append(
            {
                "sequence_pk": sequence_pk,
                "special_review": "YES" if sequence_pk in SPECIAL_REVIEW_IDS else "",
                "special_review_reason": SPECIAL_REVIEW_IDS.get(sequence_pk, ""),
                "pi_hat_name": row.get("piHatName"),
                "jonathan_sequence_id": row.get("jonathanSequenceId"),
                "staging_sample_id": row.get("stagingSampleId"),
                "staging_sample_id_2": row.get("stagingSampleId2"),
                "staging_name": row.get("stagingName"),
                "staging_hamer_name": row.get("stagingHamerName"),
                "staging_mprf_catalog_id": row.get("stagingMprfCatalogId"),
                "staging_hamer_catalog_id": row.get("stagingCatalogId"),
                "staging_biopsy_id": row.get("stagingBiopsyId"),
                "final_biopsy_id": row.get("finalBiopsyId"),
                "final_biopsy_catalog_id": row.get("finalBiopsyCatalogId"),
                "final_biopsy_sample_id": row.get("finalBiopsySampleId"),
                "final_biopsy_sequence_id": row.get("finalBiopsySequenceId"),
                "matrix_scope": row.get("matrixScope"),
                "matrix_rank": row.get("matrixRank"),
                "matrix_name": row.get("matrixName"),
                "matrix_sample_id": row.get("matrixSampleId"),
                "matrix_hamer_catalog_id": row.get("matrixHamerCatalogId"),
                "matrix_mprf_catalog_id": row.get("matrixMprfCatalogId"),
                "matrix_minimum_age": row.get("matrixMinimumAge"),
                "mapping_method": row.get("mappingMethod"),
            }
        )
    return output


def build_pair_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        a = clean(row.get("sampleIdA"))
        b = clean(row.get("sampleIdB"))
        output.append(
            {
                "pair_pk": pair_pk(a, b),
                "directional_pair_pk": f"{a}->{b}",
                "special_review_pair": "YES" if a in SPECIAL_REVIEW_IDS or b in SPECIAL_REVIEW_IDS else "",
                "sequence_pk_a": a,
                "pi_hat_name_a": row.get("nameA"),
                "matrix_name_a": row.get("matrixNameA"),
                "matrix_sample_id_a": row.get("matrixSampleIdA"),
                "jonathan_sequence_id_a": row.get("jonathanSequenceIdA"),
                "hamer_catalog_id_a": row.get("hamerCatalogIdA"),
                "mprf_catalog_id_a": row.get("mprfCatalogIdA"),
                "minimum_age_a": row.get("minimumAgeA"),
                "mapping_status_a": row.get("mappingStatusA"),
                "mapping_method_a": row.get("mappingMethodA"),
                "sequence_pk_b": b,
                "pi_hat_name_b": row.get("nameB"),
                "matrix_name_b": row.get("matrixNameB"),
                "matrix_sample_id_b": row.get("matrixSampleIdB"),
                "jonathan_sequence_id_b": row.get("jonathanSequenceIdB"),
                "hamer_catalog_id_b": row.get("hamerCatalogIdB"),
                "mprf_catalog_id_b": row.get("mprfCatalogIdB"),
                "minimum_age_b": row.get("minimumAgeB"),
                "mapping_status_b": row.get("mappingStatusB"),
                "mapping_method_b": row.get("mappingMethodB"),
                "pi_hat": row.get("PI_HAT"),
                "comparison_matrix_scope": row.get("comparisonMatrixScope"),
                "pcsu_code_a_to_b": row.get("codeAtoB"),
                "pcsu_code_b_to_a": row.get("codeBtoA"),
                "generation_prediction": row.get("generationPrediction"),
                "alignment_category": row.get("alignmentCategory"),
                "interpretation": row.get("interpretation"),
            }
        )
    return output


def build_wide_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        sequence_pk = clean(row.get("rowPiHatSampleId"))
        fixed = {
            "row_sequence_pk": sequence_pk,
            "row_special_review": "YES" if sequence_pk in SPECIAL_REVIEW_IDS else "",
            "row_pi_hat_name": row.get("rowPiHatName"),
            "row_jonathan_sequence_id": row.get("rowJonathanSequenceId"),
            "row_matrix_scope": row.get("rowMatrixScope"),
            "row_matrix_rank": row.get("rowMatrixRank"),
            "row_matrix_name": row.get("rowMatrixName"),
            "row_matrix_sample_id": row.get("rowMatrixSampleId"),
            "row_hamer_catalog_id": row.get("rowHamerCatalogId"),
            "row_mprf_catalog_id": row.get("rowMprfCatalogId"),
            "row_minimum_age": row.get("rowMinimumAge"),
        }
        matrix_values = {
            key: value
            for key, value in row.items()
            if key not in {
                "rowPiHatSampleId",
                "rowPiHatName",
                "rowJonathanSequenceId",
                "rowMatrixScope",
                "rowMatrixRank",
                "rowMatrixName",
                "rowMatrixSampleId",
                "rowHamerCatalogId",
                "rowMprfCatalogId",
                "rowMinimumAge",
            }
        }
        output.append({**fixed, **matrix_values})
    return output


def build_long_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        row_pk = clean(row.get("rowPiHatSampleId"))
        col_pk = clean(row.get("columnPiHatSampleId"))
        output.append(
            {
                "matrix_pair_pk": pair_pk(row_pk, col_pk),
                "directional_matrix_pair_pk": f"{row_pk}->{col_pk}",
                "row_sequence_pk": row_pk,
                "row_special_review": "YES" if row_pk in SPECIAL_REVIEW_IDS else "",
                "row_pi_hat_name": row.get("rowPiHatName"),
                "row_matrix_name": row.get("rowMatrixName"),
                "row_matrix_sample_id": row.get("rowMatrixSampleId"),
                "row_hamer_catalog_id": row.get("rowHamerCatalogId"),
                "row_mprf_catalog_id": row.get("rowMprfCatalogId"),
                "row_minimum_age": row.get("rowMinimumAge"),
                "column_sequence_pk": col_pk,
                "column_special_review": "YES" if col_pk in SPECIAL_REVIEW_IDS else "",
                "column_pi_hat_name": row.get("columnPiHatName"),
                "column_matrix_name": row.get("columnMatrixName"),
                "column_matrix_sample_id": row.get("columnMatrixSampleId"),
                "column_hamer_catalog_id": row.get("columnHamerCatalogId"),
                "column_mprf_catalog_id": row.get("columnMprfCatalogId"),
                "column_minimum_age": row.get("columnMinimumAge"),
                "pcsu_code": row.get("pcsuCode"),
            }
        )
    return output


def build_special_review_rows(strong_pairs: list[dict[str, Any]], sequence_by_id: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for sequence_pk, reason in SPECIAL_REVIEW_IDS.items():
        related_pairs = [
            row
            for row in strong_pairs
            if clean(row.get("sampleIdA")) == sequence_pk or clean(row.get("sampleIdB")) == sequence_pk
        ]
        source = sequence_by_id.get(sequence_pk, {})
        rows.append(
            {
                "sequence_pk": sequence_pk,
                "pi_hat_name": source.get("piHatName") or next((peer_name(row, sequence_pk) for row in related_pairs), ""),
                "review_reason": reason,
                "matrix_scope": source.get("matrixScope"),
                "matrix_name": source.get("matrixName"),
                "matrix_sample_id": source.get("matrixSampleId"),
                "hamer_catalog_id": source.get("matrixHamerCatalogId"),
                "mprf_catalog_id": source.get("matrixMprfCatalogId"),
                "jonathan_sequence_id": source.get("jonathanSequenceId"),
                "mapping_method": source.get("mappingMethod"),
                "strong_pair_count": len(related_pairs),
                "max_pi_hat_in_strong_review_pairs": max((float(row.get("PI_HAT") or 0) for row in related_pairs), default=None),
                "strong_pair_partners": "; ".join(pair_summary(row, sequence_pk) for row in related_pairs),
            }
        )
    return rows


def peer_name(row: dict[str, Any], sequence_pk: str) -> str:
    if clean(row.get("sampleIdA")) == sequence_pk:
        return clean(row.get("nameA"))
    return clean(row.get("nameB"))


def pair_summary(row: dict[str, Any], sequence_pk: str) -> str:
    if clean(row.get("sampleIdA")) == sequence_pk:
        partner = f"{row.get('sampleIdB')} {row.get('nameB')}"
    else:
        partner = f"{row.get('sampleIdA')} {row.get('nameA')}"
    return f"{partner} (PI_HAT {row.get('PI_HAT')})"


def write_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    headers = list(rows[0].keys()) if rows else ["note"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def style_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    special_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value == "YES":
                cell.fill = special_fill
    for idx, column in enumerate(ws.columns, start=1):
        max_len = min(
            55,
            max((len(str(cell.value)) if cell.value is not None else 0 for cell in column), default=8) + 2,
        )
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len)


def highlight_special_review(ws: openpyxl.worksheet.worksheet.Worksheet, *columns: str) -> None:
    headers = [cell.value for cell in ws[1]]
    column_indexes = [headers.index(column) + 1 for column in columns if column in headers]
    fill = PatternFill("solid", fgColor="FCE4D6")
    for row in range(2, ws.max_row + 1):
        values = {clean(ws.cell(row=row, column=index).value) for index in column_indexes}
        if values & SPECIAL_REVIEW_IDS.keys():
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


def pair_pk(a: str, b: str) -> str:
    return "<->".join(sorted([clean(a), clean(b)]))


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


if __name__ == "__main__":
    main()
