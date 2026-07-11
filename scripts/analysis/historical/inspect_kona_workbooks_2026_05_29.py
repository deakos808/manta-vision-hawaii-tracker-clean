import argparse
import os
import openpyxl

FILES = [
    "kona_model1_high_confidence_age_rank.xlsx",
    "kona_model2_mprf_age_classes_age_rank.xlsx",
    "kona_model3_pup_birth_year_age_rank.xlsx",
]

parser = argparse.ArgumentParser(
    description="Inspect the three workbooks used by the historical May 29, 2026 Kona comparison."
)
parser.add_argument("--input-dir", required=True, help="Directory containing the three historical model workbooks.")
args = parser.parse_args()
base = os.path.abspath(args.input_dir)

for file_name in FILES:
    workbook_path = os.path.join(base, file_name)
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    print("FILE", file_name, wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [ws.cell(1, col).value for col in range(1, min(ws.max_column, 25) + 1)]
        print(" ", sheet_name, ws.max_row, ws.max_column, headers)
